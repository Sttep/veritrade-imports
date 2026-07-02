"""
auditor_vs_app.py
─────────────────
Audita el reporte anual de la app (referencia externa) y compara
contra los outputs del pipeline de Fase 1.

MODO 1 — Solo referencia (siempre corre):
  • Analiza familia Sinotruk: variantes, consolidación, % mercado
  • Detecta duplicados/typos (SINOTRUCK HOWO vs SINOTRUK HOWO, etc.)
  • Identifica marcas del reporte que no están en configuracion.xlsx
  • Posición Sinotruk en CAMIONES Y TRACTO

MODO 2 — Con pipeline (cuando existen outputs/*_fase1.xlsx):
  • Compara unidades por marca y año: pipeline vs app
  • Calcula delta y % desvío
  • Semáforo: ✅ ≤5% | ⚠ 5-15% | ❌ >15%

Inputs:
  inputs/reporte_app_anual.xlsx  ← referencia (requerido)
  configuracion.xlsx             ← mapeos Sinotruk (recomendado)
  outputs/*_fase1.xlsx           ← pipeline Fase 1 (opcional)

Output:
  outputs/auditoria_vs_app.xlsx
"""

from __future__ import annotations
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── RUTAS ─────────────────────────────────────────────────────────────────────
REF_PATH    = Path("inputs/reporte_app_anual.xlsx")
CONFIG_PATH = Path("configuracion.xlsx")
OUTPUT_PATH = Path("outputs/auditoria_vs_app.xlsx")
PIPELINE_DIR = Path("outputs")

# ── PALETA ────────────────────────────────────────────────────────────────────
C_NEGRO   = "1A1A1A"
C_AMARILLO = "F5C400"
C_VERDE   = "1A7A40"
C_ROJO    = "B03020"
C_NARANJA = "D06010"
C_GRIS1   = "F5F3EC"
C_GRIS2   = "E8E6DC"
C_AZUL    = "1A4A7A"

SEMAFORO = {"ok": C_VERDE, "warn": C_NARANJA, "error": C_ROJO}

# ── DETECCIÓN DE AÑOS ─────────────────────────────────────────────────────────

def detectar_años(df: pd.DataFrame) -> list[int]:
    return sorted(c for c in df.columns if isinstance(c, int) and 2018 <= c <= 2030)


# ── CARGA CONFIG SINOTRUK ─────────────────────────────────────────────────────

SINOTRUK_FALLBACK = {
    "SINOTRUK": "SINOTRUK",
    "SINOTRUCK": "SINOTRUK",
    "HOWO": "SINOTRUK HOWO",
    "HOWO MAX": "SINOTRUK HOWO",
    "HOWO SINOTRUK": "SINOTRUK HOWO",
    "SINOTRUK HOWO": "SINOTRUK HOWO",
    "SINOTRUCK HOWO": "SINOTRUK HOWO",
    "SITRAK": "SINOTRUK SITRAK",
    "SINOTRUK SITRAK": "SINOTRUK SITRAK",
    "SITRAK C7H": "SINOTRUK SITRAK",
    "HONAN": "SINOTRUK HONAN",
    "HOMAN": "SINOTRUK HONAN",
    "SINOTRUK HONAN": "SINOTRUK HONAN",
    "WANGPAI": "SINOTRUK WANGPAI",
    "SINOTRUK WANGPAI": "SINOTRUK WANGPAI",
}

def cargar_config_sinotruk() -> dict:
    if not CONFIG_PATH.exists():
        print("  ⚠  configuracion.xlsx no encontrado — usando fallback")
        return SINOTRUK_FALLBACK
    try:
        df = pd.read_excel(CONFIG_PATH, sheet_name="sinotruk", dtype=str)
        mapa = {}
        for _, row in df.iterrows():
            k = str(row.get("clave", "")).upper().strip()
            v = str(row.get("valor", "")).upper().strip()
            if k and v and k != "NAN":
                mapa[k] = v
        print(f"  Config Sinotruk: {len(mapa)} variantes desde configuracion.xlsx")
        return mapa
    except Exception as e:
        print(f"  ⚠  Error leyendo config: {e} — usando fallback")
        return SINOTRUK_FALLBACK


def cargar_marcas_config() -> set:
    """Todas las marcas reconocidas en configuracion.xlsx (hoja marcas)."""
    if not CONFIG_PATH.exists():
        return set()
    try:
        df = pd.read_excel(CONFIG_PATH, sheet_name="marcas", dtype=str)
        col = df.columns[0]
        return set(df[col].dropna().str.upper().str.strip().tolist())
    except Exception:
        return set()


# ── CARGA REFERENCIA ──────────────────────────────────────────────────────────

def cargar_referencia() -> tuple[pd.DataFrame, pd.DataFrame]:
    df_m = pd.read_excel(REF_PATH, sheet_name="Volumen_Por_Marca")
    df_c = pd.read_excel(REF_PATH, sheet_name="Volumen_Por_Clase_AAP")
    df_m["marca"] = df_m["marca"].fillna("").astype(str).str.strip().str.upper()
    return df_m, df_c


# ── ANÁLISIS SINOTRUK ─────────────────────────────────────────────────────────

def analisis_sinotruk(
    df_marca: pd.DataFrame, mapa: dict, años: list[int]
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Retorna:
      df_variantes  — cada variante con submarca asignada
      df_consol     — consolidado por submarca
      df_no_mapeado — marcas que parecen Sinotruk pero no están en el mapa
    """
    es_sinotruk = df_marca["marca"].isin(mapa.keys())
    df_var = df_marca[es_sinotruk].copy()
    df_var["submarca"] = df_var["marca"].map(mapa)
    df_var["es_typo"] = df_var["marca"].apply(
        lambda m: m != mapa.get(m, m) and m not in {
            "SINOTRUK", "SINOTRUK HOWO", "SINOTRUK SITRAK",
            "SINOTRUK HONAN", "SINOTRUK WANGPAI"
        }
    )

    df_consol = (
        df_var.groupby("submarca")[años + ["TOTAL_GENERAL"]]
        .sum()
        .reset_index()
        .sort_values("TOTAL_GENERAL", ascending=False)
    )

    # ¿Hay marcas que contengan "sinotruk/howo/sitrak" pero NO están en el mapa?
    keywords = ["SINOTRUK", "HOWO", "SITRAK", "WANGPAI", "HONAN", "HOMAN"]
    patron = "|".join(keywords)
    sospechosas = df_marca[
        df_marca["marca"].str.contains(patron, na=False) & ~es_sinotruk
    ][["marca", "TOTAL_GENERAL"]]

    return df_var, df_consol, sospechosas


# ── GAPS EN CONFIG ────────────────────────────────────────────────────────────

def detectar_gaps_config(df_marca: pd.DataFrame, marcas_config: set) -> pd.DataFrame:
    """Marcas en el reporte que no están en configuracion.xlsx."""
    if not marcas_config:
        return pd.DataFrame(columns=["marca", "TOTAL_GENERAL", "nota"])
    mask = ~df_marca["marca"].isin(marcas_config)
    df_gap = df_marca[mask][["marca", "TOTAL_GENERAL"]].copy()
    df_gap = df_gap[df_gap["TOTAL_GENERAL"] > 0].sort_values("TOTAL_GENERAL", ascending=False)
    df_gap["nota"] = df_gap["TOTAL_GENERAL"].apply(
        lambda x: "🔴 ALTA PRIORIDAD" if x >= 500
        else "🟡 MEDIA" if x >= 100
        else "⚪ MENOR"
    )
    return df_gap.reset_index(drop=True)


# ── # ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN X — CARGA PIPELINE
# ══════════════════════════════════════════════════════════════════════════════

def cargar_pipeline(años: list[int]) -> pd.DataFrame | None:
    """Carga todos los *_fase1.xlsx de outputs/ (hoja 'estructurado')."""
    archivos = sorted(PIPELINE_DIR.glob("*_fase1.xlsx"))
    archivos = [a for a in archivos if "auditoria" not in a.name]
    if not archivos:
        return None

    frames = []
    for path in archivos:
        try:
            xl = pd.ExcelFile(path)
            # Fase 1 siempre usa hoja "estructurado"
            if "estructurado" not in xl.sheet_names:
                print(f"    ⚠  Sin hoja 'estructurado': {path.name}")
                continue
            df = pd.read_excel(path, sheet_name="estructurado", dtype=str)
            df["_source"] = path.name
            frames.append(df)
            print(f"    ✅ {path.name}  ({len(df):,} filas)")
        except Exception as e:
            print(f"    ⚠  No se pudo leer {path.name}: {e}")

    if not frames:
        return None

    df = pd.concat(frames, ignore_index=True)

    # ── Deduplicar entre archivos ─────────────────────────────────────────────
    df["_dedup_key"] = df.apply(
        lambda r: f"{r.get('dua_dam','')}|{r.get('vin') or r.get('chasis') or ''}",
        axis=1
    )
    antes = len(df)
    df = df.drop_duplicates(subset="_dedup_key").drop(columns="_dedup_key")
    print(f"  Registros tras deduplicar: {len(df):,} (eliminados: {antes - len(df):,})")
    
    return df


def comparar_pipeline(
    df_ref: pd.DataFrame, df_pip: pd.DataFrame | None, años: list[int]
) -> pd.DataFrame | None:
    if df_pip is None:
        return None

    # Columnas de marca y año en el output de Fase 1
    col_marca = next(
        (c for c in ["marca_normalizada", "marca_declarada", "marca"]
         if c in df_pip.columns), None
    )
    col_año = next(
        (c for c in ["año_dua", "anio_dua", "año", "anio"]
         if c in df_pip.columns), None
    )
    if not col_marca or not col_año:
        print("  ⚠  Pipeline sin columnas marca/año reconocibles")
        print(f"     Columnas disponibles: {df_pip.columns.tolist()[:10]}")
        return None

    print(f"  Columnas usadas: marca='{col_marca}' | año='{col_año}'")

    df_pip = df_pip.copy()
    df_pip["_marca"] = df_pip[col_marca].fillna("").str.upper().str.strip()
    df_pip["_año"]   = pd.to_numeric(df_pip[col_año], errors="coerce")
    df_pip = df_pip[df_pip["_año"].isin(años)]

    pivot_pip = (
        df_pip.groupby(["_marca", "_año"])
        .size()
        .reset_index(name="n")
        .pivot(index="_marca", columns="_año", values="n")
        .fillna(0)
        .astype(int)
    )
    pivot_pip.columns = [f"pip_{c}" for c in pivot_pip.columns]
    pivot_pip["pip_total"] = pivot_pip.sum(axis=1)

    pivot_ref = df_ref.set_index("marca")[años + ["TOTAL_GENERAL"]].copy()
    pivot_ref.columns = [f"app_{c}" for c in pivot_ref.columns]
    pivot_ref.rename(columns={"app_TOTAL_GENERAL": "app_total"}, inplace=True)

    df_comp = pivot_ref.join(pivot_pip, how="outer").fillna(0).reset_index()
    df_comp.rename(columns={"index": "marca"}, inplace=True)

    if "app_total" not in df_comp.columns:
        cols_app = [f"app_{a}" for a in años if f"app_{a}" in df_comp.columns]
        df_comp["app_total"] = df_comp[cols_app].sum(axis=1)

    df_comp["delta"]      = df_comp["pip_total"] - df_comp["app_total"]
    df_comp["pct_desvio"] = df_comp.apply(
        lambda r: 0 if r["app_total"] == 0
        else abs(r["delta"]) / r["app_total"] * 100, axis=1
    )
    df_comp["semaforo"] = df_comp["pct_desvio"].apply(
        lambda x: "✅ OK" if x <= 5 else "⚠ REVISAR" if x <= 15 else "❌ CRITICO"
    )
    return df_comp.sort_values("app_total", ascending=False).reset_index(drop=True)
# ── FORMATEO EXCEL ────────────────────────────────────────────────────────────

def _h(cell, bg=C_NEGRO, fg="FFFFFF"):
    cell.font = Font(bold=True, color=fg, name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _d(cell, bg=None, bold=False, halign="left"):
    cell.font = Font(name="Arial", size=9, bold=bold)
    cell.alignment = Alignment(vertical="center", horizontal=halign, wrap_text=False)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)

def _col(ws, letter, width):
    ws.column_dimensions[letter].width = width

def _freeze(ws):
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28

def formatear_libro(path: Path, df_var, df_consol, df_gap, df_clase,
                    df_resumen, df_ref_full, df_comp, años):
    wb = load_workbook(path)

    config_hojas = {
        "Resumen": (C_NEGRO, "FFFFFF"),
        "Sinotruk Variantes": (C_AMARILLO, C_NEGRO),
        "Sinotruk Consolidado": ("E6A000", C_NEGRO),
        "Marcas Sin Config": (C_ROJO, "FFFFFF"),
        "Clases AAP": ("444444", "FFFFFF"),
        "Marcas Completo": ("333333", "FFFFFF"),
        "Pipeline vs App": (C_VERDE, "FFFFFF"),
    }

    for nombre, (tab, _) in config_hojas.items():
        if nombre not in wb.sheetnames:
            continue
        ws = wb[nombre]
        ws.sheet_properties.tabColor = tab
        fg_header = _ if _ == "FFFFFF" else C_NEGRO

        for cell in ws[1]:
            if cell.value is not None:
                _h(cell, bg=tab if tab != "E6A000" else C_AMARILLO, fg=fg_header)

        for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
            bg = C_GRIS1 if r_idx % 2 == 0 else C_GRIS2
            for cell in row:
                if cell.value is None:
                    continue
                ha = "center" if isinstance(cell.value, (int, float)) else "left"
                _d(cell, bg=bg, halign=ha)

                # Semáforo en Pipeline vs App
                if nombre == "Pipeline vs App" and cell.column == ws.max_column:
                    color = C_VERDE if "OK" in str(cell.value) \
                        else C_NARANJA if "REVISAR" in str(cell.value) \
                        else C_ROJO if "CRITICO" in str(cell.value) else None
                    if color:
                        cell.font = Font(name="Arial", size=9, bold=True, color=color)

                # Prioridad en gaps
                if nombre == "Marcas Sin Config" and "ALTA" in str(cell.value):
                    cell.font = Font(name="Arial", size=9, bold=True, color=C_ROJO)

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 28

    wb.save(path)


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    print("\n" + "═" * 60)
    print("  AUDITOR: Reporte App vs Pipeline")
    print("═" * 60)

    if not REF_PATH.exists():
        print(f"\n  ❌ No encontrado: {REF_PATH}")
        print("     Coloca el archivo en inputs/reporte_app_anual.xlsx")
        return 1

    # 1. Config
    mapa_sin = cargar_config_sinotruk()
    marcas_config = cargar_marcas_config()

    # 2. Referencia
    df_marca, df_clase = cargar_referencia()
    años = detectar_años(df_marca)
    print(f"\n  Referencia: {len(df_marca)} marcas | {df_marca['TOTAL_GENERAL'].sum():,} unidades | años {años}")

    # 3. Sinotruk
    df_var, df_consol, df_sosp = analisis_sinotruk(df_marca, mapa_sin, años)
    total_sin = int(df_var["TOTAL_GENERAL"].sum())
    total_mkt = int(df_marca["TOTAL_GENERAL"].sum())
    print(f"\n  Familia Sinotruk → {len(df_var)} variantes | {total_sin:,} unidades | {total_sin/total_mkt*100:.2f}% del mercado total")
    for _, r in df_consol.iterrows():
        print(f"    {r['submarca']:<25} {int(r['TOTAL_GENERAL']):>5} unidades")
    typos = df_var[df_var["es_typo"]]
    if not typos.empty:
        print(f"\n  ⚠ Typos/alias detectados: {typos['marca'].tolist()}")
    if not df_sosp.empty:
        print(f"\n  ⚠ Marcas sospechosas NO mapeadas: {df_sosp['marca'].tolist()}")

    # 4. Gaps config
    df_gap = detectar_gaps_config(df_marca, marcas_config)
    alta_prio = df_gap[df_gap["nota"].str.contains("ALTA", na=False)]
    if not alta_prio.empty:
        print(f"\n  ❌ {len(alta_prio)} marcas alta prioridad SIN config: {alta_prio['marca'].tolist()}")

    # 5. Pipeline (opcional)
    print("\n  Buscando outputs del pipeline...")
    df_pip = cargar_pipeline(años)
    df_comp = None
    if df_pip is not None:
        print(f"  Pipeline: {len(df_pip):,} registros")
        df_comp = comparar_pipeline(df_marca, df_pip, años)
    else:
        print("  Pipeline: no encontrado (solo análisis de referencia)")

    # 6. Resumen ejecutivo
    camiones_row = df_clase[df_clase.iloc[:, 0].str.upper().str.contains("CAMION", na=False)]
    total_camiones = int(camiones_row["TOTAL_GENERAL"].sum()) if not camiones_row.empty else 0

    resumen_rows = [
        ["TOTAL UNIDADES (mercado)", *[int(df_marca[a].sum()) for a in años], total_mkt],
        ["TOTAL MARCAS DISTINTAS", *[int((df_marca[a] > 0).sum()) for a in años],
         int((df_marca["TOTAL_GENERAL"] > 0).sum())],
        ["─────────────────────", *[""] * (len(años) + 1)],
        ["SINOTRUK (consolidado)", *[int(df_var[a].sum()) for a in años], total_sin],
        ["% Sinotruk del mercado total",
         *[f"{df_var[a].sum()/df_marca[a].sum()*100:.1f}%" for a in años],
         f"{total_sin/total_mkt*100:.2f}%"],
        ["Variantes Sinotruk detectadas", *[""] * len(años), len(df_var)],
        ["Typos/alias Sinotruk", *[""] * len(años), len(typos)],
        ["─────────────────────", *[""] * (len(años) + 1)],
        ["CAMIONES Y TRACTO (clase AAP)", *[""] * len(años), total_camiones],
        ["% Sinotruk en Camiones",
         *[""] * len(años),
         f"{total_sin/total_camiones*100:.1f}%" if total_camiones else "—"],
        ["─────────────────────", *[""] * (len(años) + 1)],
        ["Marcas sin configuracion.xlsx", *[""] * len(años), len(df_gap)],
        ["Marcas alta prioridad sin config", *[""] * len(años), len(alta_prio)],
    ]
    cols_resumen = ["Métrica"] + [str(a) for a in años] + ["TOTAL"]
    df_resumen = pd.DataFrame(resumen_rows, columns=cols_resumen)

    # 7. Preparar df_var para exportar (solo columnas útiles)
    df_var_export = df_var[["marca", "submarca", "es_typo"] + años + ["TOTAL_GENERAL"]].copy()
    df_var_export["es_typo"] = df_var_export["es_typo"].map({True: "⚠ SÍ", False: "OK"})

    # 8. Escribir Excel
    OUTPUT_PATH.parent.mkdir(exist_ok=True)
    sheets = {
        "Resumen":              df_resumen,
        "Sinotruk Variantes":  df_var_export,
        "Sinotruk Consolidado": df_consol,
        "Marcas Sin Config":   df_gap,
        "Clases AAP":          df_clase,
        "Marcas Completo":     df_marca[["marca"] + años + ["TOTAL_GENERAL"]],
    }
    if df_comp is not None:
        sheets["Pipeline vs App"] = df_comp

    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        for nombre, df in sheets.items():
            df.to_excel(writer, sheet_name=nombre, index=False)

    formatear_libro(
        OUTPUT_PATH, df_var_export, df_consol, df_gap, df_clase,
        df_resumen, df_marca, df_comp, años,
    )

    print(f"\n  ✅ Auditoria exportada → {OUTPUT_PATH.resolve()}")
    print(f"{'═' * 60}\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
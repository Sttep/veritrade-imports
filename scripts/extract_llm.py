#!/usr/bin/env python3
"""Fase B — extracción estructurada con LLM (DeepSeek), híbrida y normalizada.
   [VERSIÓN SIN MODO SEGURO - SIEMPRE USA EL BATCH SIZE QUE LE PONGAS]"""

import argparse, datetime as _dt, os, sys
from pathlib import Path

def load_dotenv(path=".env"):
    p = Path(path)
    if not p.exists(): return
    for line in p.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line: continue
        k, _, val = line.partition("=")
        os.environ.setdefault(k.strip(), val.strip().strip('"').strip("'"))

import openpyxl, pandas as pd
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from scripts.llm import report, sampler, validate, vocab as vocab_mod
from scripts.llm.cache import Cache, text_key

INPUTS_DIR, OUTPUTS_DIR, HEADER_ROW = "inputs", "outputs", 6
VOLQUETE_PESADO_MIN_KG = 25000

def es_baja_confianza(row):
    try:
        if pd.isna(row.get('modelo')): return True
        if pd.isna(row.get('carroceria')): return True
        if pd.isna(row.get('marca')): return True
        km = str(row.get('kilometraje', '')).upper()
        if km in ['', 'NAN', 'NO PRECISADO', 'NO INDICADO', 'NO ESPECIFICADO']: return True
        if pd.isna(row.get('combustible')): return True
    except Exception: return True
    return False

def load_v1_with_desc(raw_path, v1_path):
    df = pd.read_excel(v1_path, sheet_name="estructurado")
    wb = openpyxl.load_workbook(raw_path, read_only=True)
    ws = wb.active
    header = next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))
    j_dua = header.index("DUA / DAM")
    j_desc = header.index("Descripcion Comercial")
    raw_data = []
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        if all(c is None for c in row): continue
        raw_data.append({"dua_dam": str(row[j_dua]).strip() if row[j_dua] else "",
                         "_desc": str(row[j_desc]).strip() if row[j_desc] else ""})
    wb.close()
    df_raw = pd.DataFrame(raw_data)
    df = df.merge(df_raw, on="dua_dam", how="left")
    print(f"  v1={len(df)} filas, crudo={len(df_raw)} filas (merge OK)")
    df["row_key"] = df.apply(lambda r: f"{r['dua_dam']}|{r['vin'] if pd.notna(r['vin']) else r['chasis']}", axis=1)
    return df

def make_on_batch():
    def on_batch(content, batch, keymap, cache):
        parsed = validate.parse_json_lenient(content)
        items = validate.items_by_index(parsed)
        for i, (_, desc) in enumerate(batch):
            item = items.get(i)
            if item is not None:
                cache.put(keymap[i], item)
    return on_batch

def obtener_marcas_desde_vocab(v):
    try:
        if hasattr(v, 'get_marcas'): return set(v.get_marcas())
        elif hasattr(v, 'brands'): return set(v.brands)
        elif hasattr(v, 'marcas'): return set(v.marcas)
        elif hasattr(v, 'vocab') and isinstance(v.vocab, dict): return set(v.vocab.keys())
        elif isinstance(v, (set, list)): return set(v)
    except Exception: pass
    return set()

def aplicar_correcciones_post_llm(df, v):
    print("  Aplicando correcciones post-LLM...")
    correcciones = []
    
    if 'carroceria' in df.columns and 'carga_util' in df.columns:
        tracto_mask = df['carroceria'] == 'TRACTOCAMIÓN'
        if tracto_mask.any():
            df.loc[tracto_mask, 'carga_util'] = 0
            correcciones.append(f"carga_util=0 para {tracto_mask.sum()} tractocamiones")
    
    if 'carroceria' in df.columns and 'kg_bruto' in df.columns:
        volq_liviano_mask = (df['carroceria'] == 'VOLQUETE') & (df['kg_bruto'] < VOLQUETE_PESADO_MIN_KG)
        if volq_liviano_mask.any():
            if 'modelo_flag' in df.columns:
                df.loc[volq_liviano_mask, 'modelo_flag'] = 'low'
                if 'flag_razon' not in df.columns:
                    df['flag_razon'] = None
                df.loc[volq_liviano_mask, 'flag_razon'] = 'volquete_liviano'
            correcciones.append(f"{volq_liviano_mask.sum()} volquetes livianos marcados como 'low'")
    
    marcas_conocidas = obtener_marcas_desde_vocab(v)
    if marcas_conocidas and 'carroceria' in df.columns:
        mascara_marca = df['carroceria'].notna() & df['carroceria'].str.upper().isin([m.upper() for m in marcas_conocidas])
        if mascara_marca.any():
            df.loc[mascara_marca, 'carroceria'] = None
            correcciones.append(f"{mascara_marca.sum()} carrocerías que eran marcas limpiadas")
    
    if 'carroceria' in df.columns and 'marca_norm' in df.columns:
        mascara_igual = df['carroceria'].notna() & df['marca_norm'].notna() & (df['carroceria'].str.upper() == df['marca_norm'].str.upper())
        if mascara_igual.any():
            df.loc[mascara_igual, 'carroceria'] = None
            correcciones.append(f"{mascara_igual.sum()} carrocerías iguales a marca_norm limpiadas")
    
    if correcciones:
        print(f"    ✅ {', '.join(correcciones)}")
    else:
        print("    ✅ Sin correcciones necesarias")
    
    return df

def process_file(raw_path, v1_path, out_path, v, cache, args):
    print(f"\n=== {raw_path.name} ===")
    df = load_v1_with_desc(raw_path, v1_path)
    
    df['_baja_confianza'] = df.apply(es_baja_confianza, axis=1)
    sub = df[df['_baja_confianza']].copy()
    print(f"  Total registros: {len(df)}")
    print(f"  Baja confianza (→ IA): {len(sub)} ({(len(sub)/len(df)*100):.1f}%)")
    print(f"  Alta confianza (skip): {len(df)-len(sub)} ({(100-len(sub)/len(df)*100):.1f}%)")
    
    use_sample = bool(args.sample) and not args.all
    if use_sample:
        sub = sampler.sample(sub, v, n=args.sample)
        print(f"  Sample: {len(sub)} registros")
    
    sub = sub.copy()
    sub["_tkey"] = sub["_desc"].map(text_key)
    pendientes = {}
    for _, r in sub.iterrows():
        k = r["_tkey"]
        if k not in cache and k not in pendientes and isinstance(r["_desc"], str):
            pendientes[k] = r["_desc"]
    print(f"  Textos pendientes: {len(pendientes)}")
    
    if args.dry_run:
        from scripts.llm.client import Stats
        stats = Stats()
    else:
        from scripts.llm.client import DeepSeekClient
        model = args.model or "deepseek-v4-flash"
        client = DeepSeekClient(
            v,
            model=model,
            batch_size=args.batch_size or 20,
            workers=args.workers or 8,
            max_tokens=2048
        )
        on_batch = make_on_batch()
        pend = dict(pendientes)
        
        # ✅ SIN MODO SEGURO - SIEMPRE USA EL BATCH SIZE QUE LE PONGAS
        if pend:
            client.batch_size = args.batch_size or 20
            client.run(list(pend.items()), cache, on_batch=on_batch)
            pend = {k: d for k, d in pendientes.items() if k not in cache}
            
        stats = client.stats
    
    recs = []
    for _, r in sub.iterrows():
        raw = cache.get(r["_tkey"])
        rec = validate.normalize_record(raw, v) if raw else validate.empty_record()
        recs.append(rec)
    
    norm_df = pd.DataFrame(recs, index=sub.index)
    out = pd.concat([sub.drop(columns=["_desc", "_tkey", "_baja_confianza"]), norm_df], axis=1)
    out = out.reset_index(drop=True)
    
    fecha = _dt.date.today().isoformat()
    modelo_usado = args.model or "deepseek-v4-flash"
    out["fuente"] = f"LLM:{modelo_usado}@{fecha}"
    
    out = aplicar_correcciones_post_llm(out, v)
    
    if 'modelo_flag' in out.columns:
        out['modelo_flag'] = out['modelo_flag'].astype(str)
    
    for col in ['marca_in_vocab', 'traccion_valido', 'combustible_valido', 'clasificacion_valido', 'caja_valido']:
        if col in out.columns:
            out[col] = out[col].astype(bool)
    
    rep = report.build(out, stats)
    
    print("\n=== REPORTE ===")
    print(rep.to_string(index=False))
    if hasattr(stats, 'prompt_tokens'):
        costo_est = (stats.prompt_tokens * 0.0000001 + stats.completion_tokens * 0.0000004)
        print(f"\n  💰 Costo estimado: ${costo_est:.6f} USD")
        print(f"  📊 Requests: {stats.requests} | Errores: {stats.errors}")
    
    revisar = pd.DataFrame()
    
    nuevos = out[(~out["marca_in_vocab"]) & out["marca_norm"].notna()]
    vocab_nuevo = (nuevos.groupby("marca_norm")
                   .agg(unidades=("marca_norm", "size"),
                        sugerencia=("marca_sugerencia", "first"),
                        modelos=("modelo_raw_llm", lambda s: ", ".join(sorted({str(x) for x in s if pd.notna(x)})[:10])))
                   .sort_values("unidades", ascending=False)
                   .reset_index())
    
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        out.to_excel(xw, sheet_name="normalizado_llm", index=False)
        if not revisar.empty:
            revisar.to_excel(xw, sheet_name="_revisar_llm", index=False)
        if not vocab_nuevo.empty:
            vocab_nuevo.to_excel(xw, sheet_name="_vocab_nuevo", index=False)
        rep.to_excel(xw, sheet_name="_reporte", index=False)
    print(f"  Escrito: {out_path} ({len(out)} filas)")
    return True

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--inputs-dir", default=INPUTS_DIR)
    ap.add_argument("--outputs-dir", default=OUTPUTS_DIR)
    ap.add_argument("--input")
    ap.add_argument("--vocab", default=None)
    ap.add_argument("--sample", type=int, default=0)
    ap.add_argument("--all", action="store_true")
    ap.add_argument("--batch-size", type=int, default=20)
    ap.add_argument("--workers", type=int, default=8)
    ap.add_argument("--model", default="deepseek-v4-flash")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    
    load_dotenv()
    v = vocab_mod.load(args.vocab) if args.vocab else vocab_mod.load()
    
    srcs = [Path(args.input)] if args.input else sorted(Path(args.inputs_dir).glob("*.xlsx"))
    if not srcs:
        print("No se encontraron .xlsx", file=sys.stderr)
        return 1
    
    out_dir = Path(args.outputs_dir)
    cache = Cache()
    ok = 0
    for raw in srcs:
        v1 = out_dir / f"{raw.stem}_estructurado.xlsx"
        if not v1.exists():
            print(f"  Falta {v1}", file=sys.stderr)
            continue
        out = out_dir / f"{raw.stem}_normalizado.xlsx"
        if process_file(raw, v1, out, v, cache, args):
            ok += 1
    
    print(f"\n{'='*60}")
    print(f"  Listo: {ok}/{len(srcs)} archivos normalizados.")
    print(f"{'='*60}")
    return 0 if ok else 1

if __name__ == "__main__":
    raise SystemExit(main())
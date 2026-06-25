"""
analisis_fase3.py — FASE 3: Inteligencia Competitiva Sinotruk
══════════════════════════════════════════════════════════════
Lee los outputs de Fase 1 y construye un reporte ejecutivo del
ecosistema Sinotruk en Perú: importadores, sub-marcas, modelos,
evolución anual y posición de Withmory.

Inputs:
  outputs/*_fase1.xlsx        ← pipeline procesado
  configuracion.xlsx          ← mapeos sinotruk + importadores

Output:
  outputs/fase3_sinotruk.xlsx (6 hojas)
    1. Resumen           — panorama del mercado Sinotruk
    2. Importadores      — detalle por dealer
    3. Marca Declarada   — HOWO vs SINOTRUK vs SITRAK por importador ★
    4. Withmory          — posición propia
    5. Evolución Anual   — año a año todos los importadores
    6. Por Carrocería    — volquete / tracto / furgón por dealer

Uso:
  python analisis_fase3.py
"""

from __future__ import annotations
from pathlib import Path
import warnings
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ── RUTAS ──────────────────────────────────────────────────────────────────────
PIPELINE_DIR = Path("outputs")
CONFIG_PATH  = Path("configuracion.xlsx")
OUTPUT_PATH  = Path("outputs/fase3_sinotruk.xlsx")

# Withmory: nombre exacto como aparece en Veritrade
WITHMORY = "CORPORATION WITHMORY S.R.L."

# Paleta Withmory
C_NEGRO    = "1A1A1A"
C_AMARILLO = "F5C400"
C_VERDE    = "1A7A40"
C_ROJO     = "B03020"
C_GRIS1    = "F5F3EC"
C_GRIS2    = "E8E6DC"
C_AZUL     = "185FA5"
C_NARANJA  = "D06010"


# ── CARGA CONFIG ───────────────────────────────────────────────────────────────

def cargar_config() -> tuple[dict, dict]:
    """Retorna (mapa_sinotruk, mapa_importador_tipo)."""
    mapa_sin, mapa_imp = {}, {}
    if not CONFIG_PATH.exists():
        print("  ⚠  configuracion.xlsx no encontrado")
        return mapa_sin, mapa_imp
    xl = pd.ExcelFile(CONFIG_PATH)
    if "sinotruk" in xl.sheet_names:
        df = xl.parse("sinotruk", dtype=str).dropna()
        cols = df.columns.tolist()
        for _, r in df.iterrows():
            k = str(r[cols[0]]).upper().strip()
            v = str(r[cols[1]]).upper().strip()
            if k and v: mapa_sin[k] = v
    if "importadores" in xl.sheet_names:
        df = xl.parse("importadores", dtype=str).dropna(subset=["importador"])
        for _, r in df.iterrows():
            imp  = str(r.get("importador","")).upper().strip()
            tipo = str(r.get("tipo","")).upper().strip()
            if imp and tipo: mapa_imp[imp] = tipo
    print(f"  Config: {len(mapa_sin)} variantes Sinotruk · "
          f"{len(mapa_imp)} importadores clasificados")
    return mapa_sin, mapa_imp


# ── CARGA PIPELINE ─────────────────────────────────────────────────────────────

def cargar_pipeline() -> pd.DataFrame:
    archivos = sorted(PIPELINE_DIR.glob("*_fase1.xlsx"))
    archivos = [a for a in archivos if "fase3" not in a.name
                and "auditoria" not in a.name]
    if not archivos:
        raise FileNotFoundError(
            f"Sin archivos *_fase1.xlsx en {PIPELINE_DIR.resolve()}")

    frames = []
    for path in archivos:
        try:
            xl = pd.ExcelFile(path)
            if "estructurado" not in xl.sheet_names:
                continue
            df = pd.read_excel(path, sheet_name="estructurado", dtype=str)
            df["_source"] = path.name
            frames.append(df)
        except Exception as e:
            print(f"    ⚠  {path.name}: {e}")

    df = pd.concat(frames, ignore_index=True)

    # Deduplicar
    df["_key"] = df.apply(
        lambda r: f"{r.get('dua_dam','')}|"
                  f"{r.get('vin') or r.get('chasis') or str(r.get('_descripcion',''))[:40]}",
        axis=1)
    antes = len(df)
    df = df.drop_duplicates(subset="_key").drop(columns="_key")
    print(f"  Pipeline: {antes:,} → {len(df):,} registros (dedup: {antes-len(df):,})")
    return df


# ── FILTRO SINOTRUK ────────────────────────────────────────────────────────────

def filtrar_sinotruk(df: pd.DataFrame, mapa_sin: dict,
                     mapa_imp: dict) -> pd.DataFrame:
    """Extrae solo vehículos de la familia Sinotruk."""
    # Buscar por submarca_sinotruk (ya procesado en Fase 1)
    mask_sub = df["submarca_sinotruk"].notna() & \
               (df["submarca_sinotruk"].str.strip() != "")

    # O por marca_declarada que esté en el mapa
    if "marca_declarada" in df.columns:
        mask_marca = df["marca_declarada"].str.upper().str.strip() \
                       .isin(mapa_sin.keys())
    else:
        mask_marca = pd.Series(False, index=df.index)

    df_sin = df[mask_sub | mask_marca].copy()

    # Asegurar campos necesarios
    for col in ["submarca_sinotruk","marca_declarada","importador",
                "tipo_importador_sinotruk","año_dua","modelo",
                "carroceria_normalizada","fob_usd"]:
        if col not in df_sin.columns:
            df_sin[col] = None

    # Completar submarca si falta
    df_sin["submarca_sinotruk"] = df_sin.apply(
        lambda r: r["submarca_sinotruk"] if pd.notna(r["submarca_sinotruk"])
                  and str(r["submarca_sinotruk"]).strip()
                  else mapa_sin.get(
                      str(r["marca_declarada"]).upper().strip(), "SINOTRUK"),
        axis=1)

    # Completar tipo importador si falta
    df_sin["tipo_importador_sinotruk"] = df_sin.apply(
        lambda r: r["tipo_importador_sinotruk"]
                  if pd.notna(r["tipo_importador_sinotruk"])
                  and str(r["tipo_importador_sinotruk"]).strip()
                  else mapa_imp.get(
                      str(r["importador"]).upper().strip(), "NO CLASIFICADO"),
        axis=1)

    df_sin["año_dua"] = pd.to_numeric(df_sin["año_dua"], errors="coerce")
    df_sin["fob_usd"] = pd.to_numeric(df_sin["fob_usd"], errors="coerce")
    df_sin["importador_upper"] = df_sin["importador"].str.upper().str.strip()

    años = sorted(df_sin["año_dua"].dropna().unique().astype(int))
    print(f"  Familia Sinotruk: {len(df_sin):,} registros | años {años}")
    return df_sin, años


# ── ANÁLISIS ───────────────────────────────────────────────────────────────────

def resumen_mercado(df: pd.DataFrame, años: list) -> pd.DataFrame:
    """Tabla resumen ejecutivo del mercado Sinotruk."""
    total = len(df)
    withmory_n = (df["importador_upper"] == WITHMORY.upper()).sum()
    oficial_n  = (df["tipo_importador_sinotruk"] == "OFICIAL").sum()
    no_ofic_n  = (df["tipo_importador_sinotruk"] == "NO OFICIAL").sum()
    directo_n  = (df["tipo_importador_sinotruk"] == "DIRECTO").sum()

    rows = [
        ["TOTAL FAMILIA SINOTRUK",
         *[int((df["año_dua"]==a).sum()) for a in años], total],
        ["── WITHMORY",
         *[int(((df["año_dua"]==a) & (df["importador_upper"]==WITHMORY.upper())).sum())
           for a in años],
         withmory_n],
        ["── OFICIAL (excl. Withmory)",
         *[int(((df["año_dua"]==a)
                & (df["tipo_importador_sinotruk"]=="OFICIAL")
                & (df["importador_upper"]!=WITHMORY.upper())).sum())
           for a in años],
         oficial_n - withmory_n],
        ["── NO OFICIAL",
         *[int(((df["año_dua"]==a)
                & (df["tipo_importador_sinotruk"]=="NO OFICIAL")).sum())
           for a in años],
         no_ofic_n],
        ["── DIRECTO / OTROS",
         *[int(((df["año_dua"]==a)
                & (df["tipo_importador_sinotruk"].isin(["DIRECTO","NO CLASIFICADO"]))).sum())
           for a in años],
         directo_n + (df["tipo_importador_sinotruk"]=="NO CLASIFICADO").sum()],
        ["─────────────────", *[""] * (len(años)+1)],
        ["SINOTRUK (marca pura)",
         *[int(((df["año_dua"]==a)
                & (df["submarca_sinotruk"]=="SINOTRUK")).sum()) for a in años],
         int((df["submarca_sinotruk"]=="SINOTRUK").sum())],
        ["SINOTRUK HOWO",
         *[int(((df["año_dua"]==a)
                & (df["submarca_sinotruk"]=="SINOTRUK HOWO")).sum()) for a in años],
         int((df["submarca_sinotruk"]=="SINOTRUK HOWO").sum())],
        ["SINOTRUK SITRAK",
         *[int(((df["año_dua"]==a)
                & (df["submarca_sinotruk"]=="SINOTRUK SITRAK")).sum()) for a in años],
         int((df["submarca_sinotruk"]=="SINOTRUK SITRAK").sum())],
        ["Otras sub-marcas",
         *[int(((df["año_dua"]==a)
                & ~df["submarca_sinotruk"].isin(
                    ["SINOTRUK","SINOTRUK HOWO","SINOTRUK SITRAK"])).sum())
           for a in años],
         int((~df["submarca_sinotruk"].isin(
             ["SINOTRUK","SINOTRUK HOWO","SINOTRUK SITRAK"])).sum())],
        ["─────────────────", *[""] * (len(años)+1)],
        ["% Withmory del total Sinotruk",
         *[f"{((df['año_dua']==a)&(df['importador_upper']==WITHMORY.upper())).sum()/(df['año_dua']==a).sum()*100:.1f}%"
           if (df['año_dua']==a).sum()>0 else "—" for a in años],
         f"{withmory_n/total*100:.1f}%"],
        ["% No Oficiales del total",
         *[f"{((df['año_dua']==a)&(df['tipo_importador_sinotruk']=='NO OFICIAL')).sum()/(df['año_dua']==a).sum()*100:.1f}%"
           if (df['año_dua']==a).sum()>0 else "—" for a in años],
         f"{no_ofic_n/total*100:.1f}%"],
    ]
    cols = ["Métrica"] + [str(a) for a in años] + ["TOTAL"]
    return pd.DataFrame(rows, columns=cols)


def tabla_importadores(df: pd.DataFrame, años: list) -> pd.DataFrame:
    """Detalle por importador: tipo, unidades por año, FOB promedio."""
    grp = df.groupby("importador")
    agg = grp.agg(
        tipo=("tipo_importador_sinotruk", "first"),
        total=("importador", "size"),
        fob_prom=("fob_usd", "mean"),
        submarcas=("submarca_sinotruk",
                   lambda x: " / ".join(sorted(x.dropna().unique()))),
        modelos_top=("modelo",
                     lambda x: ", ".join(
                         x.value_counts().head(3).index.astype(str).tolist())),
    ).reset_index()

    # Añadir columnas por año
    for año in años:
        agg[str(año)] = df[df["año_dua"]==año].groupby("importador")["importador"] \
                          .count().reindex(agg["importador"]).fillna(0).astype(int).values

    agg["fob_prom"] = agg["fob_prom"].round(0).fillna(0).astype(int)
    agg = agg.sort_values("total", ascending=False).reset_index(drop=True)

    # Columnas en orden
    cols = ["importador","tipo","total"] + \
           [str(a) for a in años] + \
           ["fob_prom","submarcas","modelos_top"]
    agg.columns.name = None
    return agg[[c for c in cols if c in agg.columns]]


def tabla_marca_declarada(df: pd.DataFrame, años: list) -> pd.DataFrame:
    """
    El insight clave: qué nombre declara cada importador en aduana.
    HOWO vs SINOTRUK vs SITRAK — posicionamiento comercial.
    """
    pivot = (df.groupby(["importador","marca_declarada","tipo_importador_sinotruk"])
               .size()
               .reset_index(name="unidades")
               .sort_values(["importador","unidades"], ascending=[True,False]))
    pivot["% del importador"] = pivot.groupby("importador")["unidades"] \
                                     .transform(lambda x: (x/x.sum()*100).round(1))
    return pivot.reset_index(drop=True)


def tabla_withmory(df: pd.DataFrame, años: list) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Posición detallada de Withmory."""
    df_w = df[df["importador_upper"] == WITHMORY.upper()].copy()

    # Por año y sub-marca
    resumen_w = df_w.groupby(["año_dua","submarca_sinotruk"]) \
                    .size().reset_index(name="unidades")
    resumen_w["año_dua"] = resumen_w["año_dua"].astype(int)

    # Market share dentro de Sinotruk por año
    total_x_año = df.groupby("año_dua").size()
    w_x_año     = df_w.groupby("año_dua").size()
    share = pd.DataFrame({
        "año": total_x_año.index.astype(int),
        "total_mercado_sinotruk": total_x_año.values,
        "withmory": w_x_año.reindex(total_x_año.index).fillna(0).astype(int).values,
    })
    share["market_share_%"] = (share["withmory"] /
                                share["total_mercado_sinotruk"] * 100).round(1)

    # Modelos top
    modelos = (df_w["modelo"].value_counts()
                              .reset_index()
                              .rename(columns={"modelo":"modelo","count":"unidades"})
                              .head(20))

    # Carrocerías
    carr = df_w["carroceria_normalizada"].value_counts() \
                                         .reset_index() \
                                         .rename(columns={
                                             "carroceria_normalizada":"carroceria",
                                             "count":"unidades"})

    df_detalle = pd.concat([
        pd.DataFrame([["── MARKET SHARE WITHMORY ──","",""]],
                     columns=["métrica","valor",""]),
        share.rename(columns={"año":"métrica",
                               "total_mercado_sinotruk":"total Sinotruk",
                               "withmory":"Withmory",
                               "market_share_%":"share %"}),
        pd.DataFrame([["","",""]]*2, columns=["métrica","total Sinotruk","Withmory"]),
    ], ignore_index=True)

    return share, modelos, carr, df_w


def tabla_evolucion(df: pd.DataFrame, años: list) -> pd.DataFrame:
    """Evolución año a año por importador (top 15)."""
    top_imp = df["importador"].value_counts().head(15).index.tolist()
    df_top  = df[df["importador"].isin(top_imp)]

    pivot = (df_top.groupby(["importador","año_dua"])
                   .size()
                   .reset_index(name="n")
                   .pivot(index="importador", columns="año_dua", values="n")
                   .fillna(0).astype(int))
    pivot.columns = [str(int(c)) for c in pivot.columns]
    pivot["TOTAL"] = pivot.sum(axis=1)

    # Añadir tipo
    tipo_map = df.drop_duplicates("importador") \
                  .set_index("importador")["tipo_importador_sinotruk"]
    pivot.insert(0, "tipo", tipo_map.reindex(pivot.index).values)

    # Marcar Withmory
    pivot.insert(0, "dealer", pivot.index)
    pivot = pivot.reset_index(drop=True)
    pivot["es_withmory"] = pivot["dealer"].str.upper().str.strip() == WITHMORY.upper()
    pivot = pivot.sort_values("TOTAL", ascending=False)
    return pivot


def tabla_carroceria(df: pd.DataFrame, años: list) -> pd.DataFrame:
    """Unidades por carrocería x importador (top 10 importadores)."""
    top_imp = df["importador"].value_counts().head(10).index.tolist()
    df_top  = df[df["importador"].isin(top_imp)]

    pivot = (df_top.groupby(["importador","carroceria_normalizada"])
                   .size()
                   .reset_index(name="n")
                   .pivot(index="importador", columns="carroceria_normalizada", values="n")
                   .fillna(0).astype(int))
    pivot["TOTAL"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("TOTAL", ascending=False)
    pivot.insert(0, "importador", pivot.index)
    return pivot.reset_index(drop=True)


# ── FORMATO EXCEL ──────────────────────────────────────────────────────────────

def _h(cell, bg=C_NEGRO, fg="FFFFFF"):
    cell.font  = Font(bold=True, color=fg, name="Arial", size=10)
    cell.fill  = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)

def _d(cell, bg=None, bold=False, halign="left", color="000000"):
    cell.font  = Font(name="Arial", size=9, bold=bold, color=color)
    cell.alignment = Alignment(vertical="center", horizontal=halign)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)

def formatear(path: Path):
    wb = load_workbook(path)
    tabs = {
        "Resumen":          (C_NEGRO,    "FFFFFF"),
        "Importadores":     ("333333",   "FFFFFF"),
        "Marca Declarada":  (C_AMARILLO, C_NEGRO),
        "Withmory":         (C_AZUL,     "FFFFFF"),
        "Evolución Anual":  ("444444",   "FFFFFF"),
        "Por Carrocería":   ("555555",   "FFFFFF"),
    }
    for nombre, (tab, fg_h) in tabs.items():
        if nombre not in wb.sheetnames: continue
        ws = wb[nombre]
        ws.sheet_properties.tabColor = tab

        for cell in ws[1]:
            if cell.value is not None:
                _h(cell, bg=tab if tab != C_AMARILLO else C_AMARILLO,
                   fg=fg_h)

        for r_idx, row in enumerate(ws.iter_rows(min_row=2), 1):
            bg = C_GRIS1 if r_idx % 2 == 0 else C_GRIS2
            for cell in row:
                if cell.value is None: continue
                ha = "center" if isinstance(cell.value, (int, float)) else "left"
                _d(cell, bg=bg, halign=ha)

                # Resaltar filas Withmory
                if nombre in ("Importadores","Evolución Anual","Marca Declarada"):
                    val = str(ws.cell(row=cell.row, column=1).value or "").upper()
                    if WITHMORY.upper()[:15] in val:
                        cell.font = Font(name="Arial", size=9, bold=True,
                                         color=C_AZUL)

                # Colores tipo en Importadores
                if nombre == "Importadores" and cell.column == 2:
                    color_map = {"OFICIAL": C_VERDE, "NO OFICIAL": C_ROJO,
                                 "DIRECTO": C_AZUL}
                    color = color_map.get(str(cell.value or "").upper(), "000000")
                    cell.font = Font(name="Arial", size=9, bold=True, color=color)

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 28
        for col in ws.columns:
            max_len = max(
                (len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    wb.save(path)


# ── MAIN ───────────────────────────────────────────────────────────────────────

def main():
    print("\n" + "═"*60)
    print("  FASE 3 — Inteligencia Competitiva Sinotruk")
    print("═"*60 + "\n")

    mapa_sin, mapa_imp = cargar_config()
    df_all = cargar_pipeline()

    df_sin, años = filtrar_sinotruk(df_all, mapa_sin, mapa_imp)

    if df_sin.empty:
        print("  ❌ Sin registros Sinotruk encontrados")
        return 1

    print("\n  Construyendo análisis...")

    df_resumen     = resumen_mercado(df_sin, años)
    df_importadores = tabla_importadores(df_sin, años)
    df_marca_dec   = tabla_marca_declarada(df_sin, años)
    share, modelos, carr, df_w = tabla_withmory(df_sin, años)
    df_evolucion   = tabla_evolucion(df_sin, años)
    df_carroceria  = tabla_carroceria(df_sin, años)

    # Print consola
    print(f"\n  Withmory market share Sinotruk:")
    for _, r in share.iterrows():
        print(f"    {int(r['año'])}: {int(r['withmory']):>4} uds "
              f"/ {int(r['total_mercado_sinotruk']):>4} total "
              f"= {r['market_share_%']:>5.1f}%")

    print(f"\n  Top importadores Sinotruk:")
    for _, r in df_importadores.head(8).iterrows():
        w = " ← WITHMORY" if WITHMORY.upper()[:15] in str(r["importador"]).upper() else ""
        print(f"    {r['importador'][:45]:<45} {int(r['total']):>5} uds  "
              f"[{r['tipo']}]{w}")

    print(f"\n  Marca declarada por importador:")
    top5_imp = df_importadores.head(5)["importador"].tolist()
    for imp in top5_imp:
        sub = df_marca_dec[df_marca_dec["importador"]==imp]
        marcas = " | ".join(
            f"{r['marca_declarada']}({int(r['unidades'])})"
            for _, r in sub.iterrows())
        print(f"    {imp[:40]:<40} → {marcas}")

    # Exportar
    OUTPUT_PATH.parent.mkdir(exist_ok=True)

    # Hoja Withmory: combinar share + modelos + carrocerías
    df_withmory_export = pd.concat([
        share.rename(columns={"año":"Año",
                               "total_mercado_sinotruk":"Total Sinotruk",
                               "withmory":"Withmory",
                               "market_share_%":"Market Share %"}),
        pd.DataFrame([["","","",""]]*2,
                     columns=["Año","Total Sinotruk","Withmory","Market Share %"]),
        pd.DataFrame([["TOP MODELOS WITHMORY","","",""]],
                     columns=["Año","Total Sinotruk","Withmory","Market Share %"]),
        modelos.rename(columns={"modelo":"Año","unidades":"Total Sinotruk"})
               .assign(**{"Withmory":"","Market Share %":""}),
    ], ignore_index=True)

    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as w:
        df_resumen.to_excel(w,       sheet_name="Resumen",         index=False)
        df_importadores.to_excel(w,  sheet_name="Importadores",    index=False)
        df_marca_dec.to_excel(w,     sheet_name="Marca Declarada", index=False)
        df_withmory_export.to_excel(w, sheet_name="Withmory",      index=False)
        df_evolucion.to_excel(w,     sheet_name="Evolución Anual", index=False)
        df_carroceria.to_excel(w,    sheet_name="Por Carrocería",  index=False)

    formatear(OUTPUT_PATH)
    print(f"\n  ✅ Guardado → {OUTPUT_PATH.resolve()}")
    print(f"{'═'*60}\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

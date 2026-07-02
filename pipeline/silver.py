"""
pipeline/silver.py  —  FASE 1 (Bronze → Silver): Extracción estructurada
═══════════════════════════════════════════════════════════════════════════

Lee configuracion.xlsx como fuente única de verdad (NO hay dicts hardcodeados).
Procesa cada .xlsx en data/bronze/ y escribe resultados estructurados en data/silver/.

Uso:
  python pipeline/silver.py
  python pipeline/silver.py --input data/bronze/mi_archivo.xlsx
"""

from __future__ import annotations
import argparse
import re
from pathlib import Path
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN 1 — RUTAS Y CONSTANTES
# ══════════════════════════════════════════════════════════════════════════════

ROOT        = Path(__file__).resolve().parent.parent
INPUTS_DIR  = ROOT / "data" / "bronze"
OUTPUTS_DIR = ROOT / "data" / "silver"
CONFIG_PATH = ROOT / "configuracion.xlsx"

# Partidas no vehiculares (autos/SUV) que a veces se cuelan en los exports de
# Veritrade filtrados por importador/rango — no son camiones, se excluyen.
PARTIDAS_EXCLUIDAS = {"8703229020", "8703210010"}

# Códigos que Veritrade usa dentro de "Descripcion Comercial"
CODES: dict[str, tuple[str, str]] = {
    "MARCA":        ("marca_declarada",  "text"),
    "MODELO":       ("modelo",           "text"),
    "VERSION":      ("version",          "text"),
    "AÑO MOD":      ("anio_modelo",      "int"),
    "AÑO":          ("anio_modelo",      "int"),
    "VI":           ("vin",              "text"),
    "VIN":          ("vin",              "text"),
    "CH":           ("chasis",           "text"),
    "MO":           ("motor_serie",      "text"),
    "CO":           ("combustible",      "text"),
    "COMB":         ("combustible",      "text"),
    "C1":           ("color",            "text"),
    "COLOR":        ("color",            "text"),
    "NC":           ("num_cilindros",    "int"),
    "CC":           ("cilindrada_cc",    "num"),
    "PM":           ("potencia",         "power"),
    "EJ":           ("ejes",             "int"),
    "FR":           ("traccion",         "text"),
    "TT":           ("transmision",      "text"),
    "CA":           ("carroceria",       "text"),
    "AS":           ("asientos",         "int"),
    "PA":           ("puertas",          "int"),
    "PB":           ("peso_bruto_desc",  "num"),
    "PN":           ("peso_neto_desc",   "num"),
    "CU":           ("carga_util",       "num"),
    "LA":           ("largo_mm",         "num"),
    "AN":           ("ancho_mm",         "num"),
    "AL":           ("alto_mm",          "num"),
    "DE":           ("dist_ejes",        "num"),
    "KILOMETRAJE":  ("kilometraje",      "num"),
    "KM":           ("kilometraje",      "num"),
    "KLM":          ("kilometraje",      "num"),
    "CLASIFICACION":("clasificacion",    "text"),
    "CAJA":         ("caja",             "text"),
    "VE":           ("version",          "text"),
    "CH/VIN":       ("vin",              "text"),
    "SN": ("_", "text"), "NR": ("_", "text"), "AR": ("_", "text"),
    "SD": ("_", "text"), "SP": ("_", "text"), "BC": ("_", "text"),
    "PP": ("_", "text"), "TE": ("_", "text"), "ACD": ("_", "text"),
    "LET": ("_", "text"), "ARA": ("_", "text"), "CLM": ("_", "text"),
    "BSA": ("_", "text"), "CRT": ("_", "text"), "CTC": ("_", "text"),
    "DAS": ("_", "text"), "EAV": ("_", "text"), "EEL": ("_", "text"),
    "PEL": ("_", "text"),
}

HARD_COLS: dict[str, str] = {
    "Partida Aduanera":         "partida",
    "Partida Arancelaria":      "partida",
    "Aduana":                   "aduana",
    "DUA / DAM":                "dua_dam",
    "Fecha":                    "fecha_dua",
    "Importador":               "importador",
    "Embarcador / Exportador":  "exportador",
    "Kg Bruto":                 "kg_bruto",
    "Kg Neto":                  "kg_neto",
    "Qty 1":                    "qty_1",
    "Und 1":                    "und_1",
    "Qty 2":                    "qty_2",
    "Und 2":                    "und_2",
    "U$ FOB Tot":               "fob_usd",
    "U$ CFR Tot":               "cfr_usd",
    "U$ CIF Tot":               "cif_usd",
    "Pais de Origen":           "pais_origen",
    "País Origen":              "pais_origen",
    "Pais de Compra":           "pais_compra",
    "Puerto de Embarque":       "puerto_embarque",
    "Via":                      "via",
    "Agente de Aduana":         "agente_aduana",
    "Estado":                   "estado",
    "Descripcion Comercial":    "_descripcion",
    "Descripcion del Producto": "_descripcion",
    "Descripción Comercial":    "_descripcion",
}

ATU = [(3500, "N1"), (12000, "N2"), (float("inf"), "N3")]

MESES = {1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
         7:"Julio",8:"Agosto",9:"Setiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}
TRIMESTRES = {m: f"Q{(m-1)//3+1}" for m in range(1,13)}

TRANSMISION_NORM = {
    "MECANICO":"Mecánico","MECÁNICO":"Mecánico","MANUAL":"Mecánico",
    "AUTOMATICO":"Automático","AUTOMÁTICO":"Automático","AUTOMATIC":"Automático",
    "AT":"Automático","MT":"Mecánico","M/T":"Mecánico","A/T":"Automático",
}

C_NEGRO = "1A1A1A"
C_AMARILLO = "F5C400"
C_VERDE = "1A7A40"
C_ROJO = "B03020"
C_GRIS1 = "F5F3EC"
C_GRIS2 = "E8E6DC"


# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN 2 — CARGA DE CONFIGURACION.XLSX
# ══════════════════════════════════════════════════════════════════════════════

class Config:
    def __init__(self, path: Path = CONFIG_PATH):
        if not path.exists():
            print(f"  ⚠  {path} no encontrado — config mínima activa")
            self._set_fallbacks()
            return
        self.marcas_set:       set[str]        = set()
        self.marca_map:        dict[str,str]    = {}
        self.brand_re:         re.Pattern       = re.compile(r"(?!)")
        self.modelos:          dict[str,set]    = {}
        self.carroceria_map:   dict[str,str]    = {}
        self.excluir_set:      set[str]         = set()
        self.sinotruk_map:     dict[str,str]    = {}
        self.importador_tipo:  dict[str,str]    = {}
        self.estados_nuevos:   set[str]         = {
            "NUEVO","NUEVA","NEW","NUEVO/0 KM","0 KM","SIN USO",
            "NUEVO SIN USO","NUEVA SIN USO","NUEVO DESARMADO","NUEVO SEMIARMADO"
        }
        self._cargar(path)
        print(f"  Config: {len(self.marcas_set)} marcas · "
              f"{len(self.carroceria_map)} carrocerías · "
              f"{len(self.sinotruk_map)} variantes Sinotruk · "
              f"{len(self.importador_tipo)} importadores clasificados")

    def _cargar(self, path: Path):
        xl = pd.ExcelFile(path)

        if "marcas" in xl.sheet_names:
            df = xl.parse("marcas", dtype=str)
            cols = df.columns.tolist()
            col_bruta = cols[0]
            col_norm  = cols[1] if len(cols) > 1 else cols[0]
            df = df.dropna(subset=[col_bruta])
            for _, row in df.iterrows():
                bruta = str(row[col_bruta]).upper().strip()
                norm_val = row[col_norm]
                norm  = str(norm_val).upper().strip() if pd.notna(norm_val) else ""
                if bruta:
                    self.marca_map[bruta] = norm or bruta
            self.marcas_set = set(self.marca_map.keys())
            brands = sorted(self.marcas_set, key=len, reverse=True)
            self.brand_re = re.compile(
                r"\b(" + "|".join(re.escape(b) for b in brands) + r")\b", re.I
            )

        if "modelos" in xl.sheet_names:
            df = xl.parse("modelos", dtype=str).dropna(subset=["marca"])
            for _, row in df.iterrows():
                m = str(row.get("marca","")).upper().strip()
                mod = str(row.get("modelo","")).upper().strip()
                if m and mod:
                    self.modelos.setdefault(m, set()).add(mod)

        if "carrocerias" in xl.sheet_names:
            df = xl.parse("carrocerias", dtype=str).dropna()
            cols = df.columns.tolist()
            for _, row in df.iterrows():
                k = str(row[cols[0]]).upper().strip()
                v = str(row[cols[1]]).upper().strip()
                if k and v:
                    self.carroceria_map[k] = v

        if "excluir" in xl.sheet_names:
            df = xl.parse("excluir", dtype=str).iloc[:, 0].dropna()
            self.excluir_set = set(df.str.upper().str.strip())

        if "sinotruk" in xl.sheet_names:
            df = xl.parse("sinotruk", dtype=str).dropna()
            cols = df.columns.tolist()
            for _, row in df.iterrows():
                k = str(row[cols[0]]).upper().strip()
                v = str(row[cols[1]]).upper().strip()
                if k and v:
                    self.sinotruk_map[k] = v

        if "importadores" in xl.sheet_names:
            df = xl.parse("importadores", dtype=str).dropna(subset=["importador"])
            for _, row in df.iterrows():
                imp = str(row.get("importador","")).upper().strip()
                tipo = str(row.get("tipo","")).upper().strip()
                if imp and tipo:
                    self.importador_tipo[imp] = tipo

    def _set_fallbacks(self):
        self.marcas_set      = set()
        self.marca_map       = {}
        self.brand_re        = re.compile(r"(?!)")
        self.modelos         = {}
        self.carroceria_map  = {}
        self.excluir_set     = set()
        self.sinotruk_map    = {"SINOTRUK":"SINOTRUK","HOWO":"SINOTRUK HOWO",
                                "SITRAK":"SINOTRUK SITRAK","HOMAN":"SINOTRUK HONAN"}
        self.importador_tipo = {}
        self.estados_nuevos  = {"NUEVO","NUEVA","NEW","0 KM","SIN USO"}


# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN 3 — DETECCIÓN DE HEADER Y LECTURA DE ARCHIVO
# ══════════════════════════════════════════════════════════════════════════════

ANCHOR_KEYS = {"DUA", "DAM", "IMPORTADOR", "DESCRIPCI", "ESTADO", "FECHA"}

def detectar_header_row(path: Path, max_scan: int = 20) -> int:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    for row_idx, row in enumerate(ws.iter_rows(max_row=max_scan, values_only=True), 1):
        vals = " ".join(str(c or "").upper() for c in row)
        if sum(1 for k in ANCHOR_KEYS if k in vals) >= 2:
            wb.close()
            return row_idx
    wb.close()
    return 1


def normalizar_str(s: str) -> str:
    return (s.lower().strip()
            .replace("á","a").replace("é","e").replace("í","i")
            .replace("ó","o").replace("ú","u").replace("ü","u"))


def mapear_columnas(df_cols: list[str]) -> dict[str, str]:
    mapa = {}
    norm_df = {normalizar_str(c): c for c in df_cols}

    for hdr_orig, interno in HARD_COLS.items():
        if interno in mapa.values():
            continue
        norm_h = normalizar_str(hdr_orig)
        if norm_h in norm_df:
            mapa[norm_df[norm_h]] = interno
            continue
        for norm_col, real_col in norm_df.items():
            if norm_h in norm_col or norm_col in norm_h:
                if real_col not in mapa:
                    mapa[real_col] = interno
                break
    return mapa


def leer_archivo(path: Path) -> tuple[pd.DataFrame, dict]:
    header_row = detectar_header_row(path)
    df = pd.read_excel(path, header=header_row - 1, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    col_map = mapear_columnas(df.columns.tolist())
    df = df.rename(columns=col_map)
    meta = {"archivo": path.name, "header_row": header_row,
            "filas_raw": len(df), "cols_mapeadas": len(col_map)}
    return df, meta


# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN 4 — PARSEO DE DESCRIPCION COMERCIAL
# ══════════════════════════════════════════════════════════════════════════════

def _build_code_pattern() -> re.Pattern:
    codes_sorted = sorted(CODES.keys(), key=len, reverse=True)
    escaped      = [re.escape(c) for c in codes_sorted]
    code_group   = "|".join(escaped)
    return re.compile(
        rf"({code_group})\s*[:=]\s*(.*?)(?=\s*[,;]?\s*(?:{code_group})\s*[:=]|$)",
        re.IGNORECASE,
    )

CODE_PATTERN = _build_code_pattern()


def _convertir(valor: str, tipo: str):
    valor = valor.strip().upper()
    if not valor:
        return None
    if tipo == "int":
        m = re.search(r"\d{4}", valor) or re.search(r"\d+", valor)
        return int(m.group()) if m else None
    if tipo == "num":
        m = re.search(r"\d+(?:[.,]\d+)?", valor)
        if not m:
            return None
        try:
            return float(m.group().rstrip(".,").replace(",", "."))
        except ValueError:
            return None
    if tipo == "power":
        m = re.search(r"([\d.,]+)\s*(HP|KW|CV|PS)", valor, re.I)
        if m:
            return f"{m.group(1).rstrip('.,')  } {m.group(2).upper()}"
        m = re.search(r"([\d.,]+)\s*@", valor)
        if m:
            return m.group(1).rstrip(".,")
        m = re.search(r"\d+(?:[.,]\d+)?", valor)
        if not m:
            return None
        try:
            return float(m.group().rstrip(".,").replace(",", "."))
        except ValueError:
            return None
    return valor.strip(" ,;") or None


def parsear_descripcion(texto: str) -> dict:
    if not texto or not isinstance(texto, str) or texto.strip() in ("nan","None",""):
        return {}
    resultado: dict = {}
    for match in CODE_PATTERN.finditer(texto):
        code  = match.group(1).strip().upper()
        valor = match.group(2).strip()
        if code in CODES:
            campo, tipo = CODES[code]
            if campo != "_" and campo not in resultado:
                converted = _convertir(valor, tipo)
                if converted is not None:
                    resultado[campo] = converted
    return resultado


# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN 5 — NORMALIZACIÓN Y CLASIFICACIÓN
# ══════════════════════════════════════════════════════════════════════════════

def normalizar_transmision(valor: str | None) -> str | None:
    if not valor:
        return None
    return TRANSMISION_NORM.get(str(valor).upper().strip(), str(valor).strip())


def clasificar_atu(peso: float | None) -> str | None:
    if peso is None:
        return None
    for limite, cat in ATU:
        if peso <= limite:
            return cat
    return "N3"


def parsear_fecha(fecha_raw) -> dict:
    resultado = {"año_dua": None, "mes_dua": None,
                 "mes_nombre": None, "trimestre_dua": None}
    try:
        if isinstance(fecha_raw, str):
            fecha_raw = fecha_raw.strip()
            m = re.match(r"(\d{4})[-/](\d{1,2})", fecha_raw)
            if m:
                año, mes = int(m.group(1)), int(m.group(2))
            else:
                m = re.match(r"(\d{1,2})[-/](\d{4})", fecha_raw)
                if m:
                    mes, año = int(m.group(1)), int(m.group(2))
                else:
                    return resultado
        elif hasattr(fecha_raw, "year"):
            año, mes = fecha_raw.year, fecha_raw.month
        else:
            return resultado

        if 2015 <= año <= 2030 and 1 <= mes <= 12:
            resultado.update({
                "año_dua": año,
                "mes_dua": mes,
                "mes_nombre": MESES[mes],
                "trimestre_dua": TRIMESTRES[mes],
            })
    except Exception:
        pass
    return resultado


def buscar_marca(texto: str, cfg: Config) -> str | None:
    if not texto:
        return None
    m = cfg.brand_re.search(texto)
    return m.group(0).upper() if m else None


def normalizar_carroceria(valor: str | None, cfg: Config) -> str | None:
    if not valor:
        return None
    v = str(valor).upper().strip()
    if v in cfg.carroceria_map:
        return cfg.carroceria_map[v]
    for k, norm in cfg.carroceria_map.items():
        if k in v or v in k:
            return norm
    return v


def debe_excluir(row: dict, cfg: Config) -> tuple[bool, str]:
    partida = str(row.get("partida") or "").strip()
    if partida in PARTIDAS_EXCLUIDAS:
        return True, f"partida_no_vehicular={partida}"

    desc = str(row.get("_descripcion") or "").upper().strip()
    m_cat = re.match(r"^(L[1-6])\s*[,\s]", desc)
    if m_cat:
        return True, f"categoria_L={m_cat.group(1)}"

    estado = str(row.get("estado") or "").upper().strip()
    if estado and estado not in cfg.estados_nuevos:
        return True, f"estado={estado}"

    # Solo excluir por km si el vehículo NO es nuevo (km de fábrica/transporte son normales)
    km = row.get("kilometraje")
    if estado not in cfg.estados_nuevos:
        try:
            if km is not None and float(km) > 200:
                return True, f"km={km}"
        except (ValueError, TypeError):
            pass

    carroceria = str(row.get("carroceria", "") or "").upper().strip()
    if carroceria in cfg.excluir_set:
        return True, f"carroceria_excluida={carroceria}"

    marca = str(row.get("marca_declarada", "") or "").upper().strip()
    if marca in cfg.excluir_set:
        return True, f"marca_excluida={marca}"

    for termino in cfg.excluir_set:
        if len(termino) > 4 and termino in desc:
            return True, f"texto_excluido={termino}"

    return False, ""


def _extraer_posicional(desc: str, cfg: Config) -> dict:
    resultado: dict = {}
    texto_up = str(desc).upper().strip()

    texto_sin_cat = re.sub(r"^(?:N[123]|M[123]|L[12456])\s*[,\s]\s*", "", texto_up).strip()
    if not texto_sin_cat:
        return resultado

    m_inicio = CODE_PATTERN.search(texto_sin_cat)
    cabecera = texto_sin_cat[:m_inicio.start()].strip() if m_inicio else texto_sin_cat

    if "," in cabecera:
        tokens = [t.strip(" ,.") for t in cabecera.split(",") if t.strip(" ,.")]
        if tokens:
            resultado["_marca_pos"] = tokens[0]
        if len(tokens) >= 2:
            resultado["_modelo_pos"] = tokens[1].split()[0]
    else:
        m = cfg.brand_re.search(cabecera)
        if m:
            resultado["_marca_pos"] = m.group(0)
            resto = cabecera[m.end():].strip()
            tokens_resto = [t for t in resto.split()
                           if len(t) > 2
                           and t not in {"CHASIS","CABINADO","CABINA","DOBLE","AÑO","THE"}
                           and not re.match(r"^\d{4}$", t)]
            if tokens_resto:
                resultado["_modelo_pos"] = tokens_resto[0]

    CARROCERIA_KW = {
        "CHASIS CABINADO": "CHASIS CABINADO",
        "CHASIS CABINA":   "CHASIS CABINADO",
        "PICK UP":         "PICK UP",
        "FURGON":          "FURGÓN",
        "FURGÓN":          "FURGÓN",
        "VOLQUETE":        "VOLQUETE",
        "CISTERNA":        "CISTERNA",
        "PLATAFORMA":      "PLATAFORMA",
        "TRACTO":          "TRACTOCAMIÓN",
        "TRACTOCAMION":    "TRACTOCAMIÓN",
    }
    for kw, norm in CARROCERIA_KW.items():
        if kw in cabecera:
            resultado["_carroceria_pos"] = norm
            break

    return resultado


# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN 6 — PROCESAMIENTO POR FILA
# ══════════════════════════════════════════════════════════════════════════════

def procesar_fila(row: pd.Series, cfg: Config) -> dict:
    desc = str(row.get("_descripcion", "") or "")

    extraido  = parsear_descripcion(desc)
    posicional = _extraer_posicional(desc, cfg)

    marca_declarada = (extraido.get("marca_declarada")
                       or posicional.get("_marca_pos")
                       or buscar_marca(desc, cfg))
    marca_normalizada = None
    if marca_declarada:
        mk_up = str(marca_declarada).upper().strip()
        if mk_up in cfg.marcas_set:
            marca_normalizada = cfg.marca_map.get(mk_up, mk_up)
        else:
            m = cfg.brand_re.search(mk_up)
            if m:
                matched = m.group(0).upper()
                marca_normalizada = cfg.marca_map.get(matched, matched)

    modelo = extraido.get("modelo") or posicional.get("_modelo_pos")

    submarca_sinotruk = None
    if marca_declarada:
        submarca_sinotruk = cfg.sinotruk_map.get(str(marca_declarada).upper().strip())

    carroceria_raw  = extraido.get("carroceria") or posicional.get("_carroceria_pos")
    carroceria_norm = normalizar_carroceria(carroceria_raw, cfg)

    transmision = normalizar_transmision(extraido.get("transmision"))

    peso_desc    = extraido.get("peso_bruto_desc")
    kg_bruto_col = _to_num(row.get("kg_bruto"))
    peso_para_atu = peso_desc if peso_desc else kg_bruto_col
    categoria_atu = clasificar_atu(peso_para_atu)

    fecha_info = parsear_fecha(row.get("fecha_dua"))

    importador = str(row.get("importador", "") or "").strip()
    tipo_imp   = cfg.importador_tipo.get(importador.upper().strip())

    confianza = "alta" if (marca_normalizada and carroceria_norm) else "baja"

    return {
        "dua_dam":              row.get("dua_dam"),
        "partida":              row.get("partida"),
        "aduana":               row.get("aduana"),
        "fecha_dua":            row.get("fecha_dua"),
        **fecha_info,
        "importador":           importador,
        "tipo_importador_sinotruk": tipo_imp,
        "exportador":           row.get("exportador"),
        "pais_origen":          row.get("pais_origen"),
        "pais_compra":          row.get("pais_compra"),
        "estado":               row.get("estado"),
        "marca_declarada":      marca_declarada,
        "marca_normalizada":    marca_normalizada,
        "submarca_sinotruk":    submarca_sinotruk,
        "modelo":               modelo,
        "version":              extraido.get("version"),
        "anio_modelo":          extraido.get("anio_modelo"),
        "carroceria":           carroceria_raw,
        "carroceria_normalizada": carroceria_norm,
        "combustible":          extraido.get("combustible"),
        "color":                extraido.get("color"),
        "traccion":             extraido.get("traccion"),
        "transmision":          transmision,
        "num_cilindros":        extraido.get("num_cilindros"),
        "cilindrada_cc":        extraido.get("cilindrada_cc"),
        "potencia":             extraido.get("potencia"),
        "ejes":                 extraido.get("ejes"),
        "asientos":             extraido.get("asientos"),
        "peso_bruto_desc":      peso_desc,
        "peso_neto_desc":       extraido.get("peso_neto_desc"),
        "carga_util":           extraido.get("carga_util"),
        "kg_bruto_col":         kg_bruto_col,
        "largo_mm":             extraido.get("largo_mm"),
        "ancho_mm":             extraido.get("ancho_mm"),
        "alto_mm":              extraido.get("alto_mm"),
        "dist_ejes":            extraido.get("dist_ejes"),
        "kilometraje":          extraido.get("kilometraje"),
        "vin":                  extraido.get("vin"),
        "chasis":               extraido.get("chasis"),
        "motor_serie":          extraido.get("motor_serie"),
        "clasificacion":        extraido.get("clasificacion"),
        "caja":                 extraido.get("caja"),
        "categoria_atu":        categoria_atu,
        "fob_usd":              _to_num(row.get("fob_usd")),
        "cif_usd":              _to_num(row.get("cif_usd")),
        "kg_neto":              _to_num(row.get("kg_neto")),
        "qty_1":                row.get("qty_1"),
        "und_1":                row.get("und_1"),
        "_descripcion":         desc,
        "confianza":            confianza,
    }


def _to_num(val) -> float | None:
    try:
        return float(str(val).replace(",", ".")) if val not in (None, "", "nan", "None") else None
    except (ValueError, TypeError):
        return None


# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN 7 — PIPELINE POR ARCHIVO
# ══════════════════════════════════════════════════════════════════════════════

def procesar_archivo(path: Path, cfg: Config, force: bool = False) -> bool:
    print(f"\n  ═══ {path.name} ═══")
    out_path = OUTPUTS_DIR / f"{path.stem}_fase1.xlsx"
    if out_path.exists() and not force:
        print("  ⏭  Ya procesado — saltando")
        return True

    df_raw, meta = leer_archivo(path)
    print(f"  Header detectado en fila {meta['header_row']} | "
          f"{meta['filas_raw']} registros | "
          f"{meta['cols_mapeadas']} columnas mapeadas")

    if "_descripcion" not in df_raw.columns:
        print("  ⚠  Sin columna 'Descripcion Comercial' — archivo omitido")
        return False

    registros_ok = []
    excluidos    = []

    filas    = df_raw.to_dict("records")
    total    = len(filas)
    intervalo = max(1000, total // 20)

    for i, row in enumerate(filas):
        if i > 0 and i % intervalo == 0:
            print(f"    ├ {i}/{total} ({i/total*100:.0f}%)...", end="\r")
        rec = procesar_fila(row, cfg)
        excluir, razon = debe_excluir(rec, cfg)
        if excluir:
            rec["razon_exclusion"] = razon
            excluidos.append(rec)
        else:
            registros_ok.append(rec)

    if total > intervalo:
        print()

    df_ok  = pd.DataFrame(registros_ok)
    df_exc = pd.DataFrame(excluidos)

    if not df_ok.empty:
        df_ok["_dedup_key"] = df_ok.apply(
            lambda r: f"{r.get('dua_dam','')}|{r.get('vin') or r.get('chasis') or str(r.get('_descripcion',''))[:40]}",
            axis=1
        )
        antes = len(df_ok)
        df_ok = df_ok.drop_duplicates(subset="_dedup_key").drop(columns="_dedup_key")
        if antes - len(df_ok) > 0:
            print(f"  🔄 Duplicados eliminados: {antes - len(df_ok)}")

    df_f2  = df_ok[df_ok["confianza"] == "baja"].copy() if not df_ok.empty else pd.DataFrame()
    n_ok  = len(df_ok)
    n_exc = len(df_exc)
    n_f2  = len(df_f2)
    pct_baja = f"{n_f2/n_ok*100:.1f}%" if n_ok else "—"

    print(f"  Válidos:          {n_ok}")
    print(f"  Excluidos:        {n_exc}")
    print(f"  → Para Fase Gold: {n_f2}  ({pct_baja})")

    if not df_ok.empty:
        marcas = df_ok["marca_normalizada"].value_counts().head(5)
        print(f"  Top marcas:       "
              f"{', '.join(f'{m}({n})' for m, n in marcas.items())}")
        sin_marca = df_ok["marca_normalizada"].isna().sum()
        if sin_marca:
            print(f"  ⚠  Sin marca reconocida: {sin_marca} registros")

    df_log = pd.DataFrame([{
        "archivo":           meta["archivo"],
        "header_row":        meta["header_row"],
        "filas_raw":         meta["filas_raw"],
        "validos":           n_ok,
        "excluidos":         n_exc,
        "para_fase_gold":    n_f2,
        "pct_baja_confianza": pct_baja,
        "cols_mapeadas":     meta["cols_mapeadas"],
        "procesado_en":      datetime.now().isoformat(timespec="seconds"),
    }])

    n_total = len(df_ok) + len(df_exc)

    if n_total <= 50_000:
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            if not df_ok.empty:
                df_ok.to_excel(writer,  sheet_name="estructurado",  index=False)
            if not df_f2.empty:
                df_f2.to_excel(writer,  sheet_name="_para_fase_gold", index=False)
            if not df_exc.empty:
                df_exc.to_excel(writer, sheet_name="_excluidos",    index=False)
            df_log.to_excel(writer,     sheet_name="_log",          index=False)
        _formatear_salida(out_path)
    else:
        print(f"  📦 {n_total:,} filas → modo streaming (xlsxwriter)")
        with pd.ExcelWriter(
            out_path,
            engine="xlsxwriter",
            engine_kwargs={"options": {"constant_memory": True}},
        ) as writer:
            if not df_ok.empty:
                df_ok.to_excel(writer,  sheet_name="estructurado",  index=False)
            if not df_f2.empty:
                df_f2.to_excel(writer,  sheet_name="_para_fase_gold", index=False)
            if not df_exc.empty:
                df_exc.to_excel(writer, sheet_name="_excluidos",    index=False)
            df_log.to_excel(writer,     sheet_name="_log",          index=False)

    print(f"  ✅ Escrito → {out_path.relative_to(ROOT)}")
    return True


def _formatear_salida(path: Path):
    wb = openpyxl.load_workbook(path)
    configs = {
        "estructurado":    (C_NEGRO,   "FFFFFF"),
        "_para_fase_gold": ("884400",  "FFFFFF"),
        "_excluidos":      (C_ROJO,    "FFFFFF"),
        "_log":            ("444444",  "FFFFFF"),
    }
    for sheet_name, (tab, _) in configs.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        ws.sheet_properties.tabColor = tab
        for cell in ws[1]:
            if cell.value is None:
                continue
            cell.font  = Font(bold=True, color="FFFFFF", name="Arial", size=9)
            cell.fill  = PatternFill("solid", start_color=tab)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for r_idx, row in enumerate(ws.iter_rows(min_row=2), 1):
            bg = C_GRIS1 if r_idx % 2 == 0 else C_GRIS2
            for cell in row:
                if cell.value is None:
                    continue
                cell.font = Font(name="Arial", size=9)
                cell.fill = PatternFill("solid", start_color=bg)
                cell.alignment = Alignment(
                    vertical="center",
                    horizontal="center" if isinstance(cell.value, (int, float)) else "left",
                )
                if str(ws.cell(1, cell.column).value or "") == "confianza":
                    if str(cell.value) == "alta":
                        cell.font = Font(name="Arial", size=9, bold=True, color=C_VERDE)
                    elif str(cell.value) == "baja":
                        cell.font = Font(name="Arial", size=9, bold=True, color=C_ROJO)
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 26
    wb.save(path)


# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN 8 — ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(description="Fase 1 (Silver) — Extractor Veritrade")
    ap.add_argument("--input", help="Archivo específico (omitir = todos en data/bronze/)")
    ap.add_argument("--force", action="store_true", help="Reprocesar aunque ya exista el silver")
    args = ap.parse_args()

    print("\n" + "═" * 60)
    print("  FASE 1 (Bronze → Silver) — Extracción estructurada")
    print("═" * 60)

    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)

    cfg = Config(CONFIG_PATH)

    archivos = [Path(args.input)] if args.input else sorted(INPUTS_DIR.glob("*.xlsx"))
    archivos = [a for a in archivos if "reporte_app" not in a.name and "auditoria" not in a.name]

    if not archivos:
        print(f"\n  ❌ Sin archivos .xlsx en {INPUTS_DIR}/")
        return 1

    print(f"\n  {len(archivos)} archivo(s) a procesar\n")
    ok = sum(procesar_archivo(a, cfg, force=args.force) for a in archivos)

    print(f"\n{'═' * 60}")
    print(f"  Listo: {ok}/{len(archivos)} archivos procesados")
    print(f"  Silver en: {OUTPUTS_DIR}/")
    print(f"{'═' * 60}\n")
    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
